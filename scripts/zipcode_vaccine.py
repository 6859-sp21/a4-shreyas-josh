# Returns all Tuesdays and Thursdays of a given year
from datetime import datetime as dt
from datetime import date 
import calendar
import requests
from pandas import read_excel
from openpyxl import load_workbook

def get_thursdays():
    years = [2021]
    c = calendar.TextCalendar(calendar.SUNDAY)
    dates = []
    for year in years:
        for m in range(1,13):
            for i in c.itermonthdays(year,m):
                if i != 0:                                      #calendar constructs months with leading zeros (days belongng to the previous month)
                    day = date(year,m,i)
                    if day.weekday() == 3: 
                        dates.append("%s-%s-%s" % (calendar.month_name[m],i,year))
                    if date(year,m,i) == dt.today().date():
                        return dates

dates = get_thursdays()

prefix = "https://www.mass.gov/doc/weekly-covid-19-municipality-vaccination-report-"
postfix = "/download"
all_dfs = []

f = open("zipcode_vaccine.csv", "w")

for d in dates:
    url = prefix + d.lower() + postfix
    resp = requests.get(url)
    content_type = resp.headers['Content-Type']
    if content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        fname = d + '.xlsx'
        with open(fname, 'wb') as r:
            r.write(resp.content)
        #my_sheet = 'Race and ethnicity - muni.' # change it to your sheet name, you can find your sheet name at the bottom left of your excel file
        wb = load_workbook(fname)
        ws = wb[wb.sheetnames[6]]
        for row in ws.iter_rows(min_row=4, max_col=10):
            row_str = d
            add_row = True
            for cell in row:
                if str(cell.value) == 'Unspecified':
                    add_row = False
                    break
                row_str += ',' + str(cell.value).strip("/'") 
            if add_row:
                row_str = row_str.strip(',')
                row_str += '\n'
                f.write(row_str)
f.close()


def prepend_line(filename, line):
    with open(filename, 'r+') as f:
        content = f.read()
        f.seek(0, 0)
        f.write(line.rstrip('\r\n') + '\n' + content)

column_labels = 'Date,ZIP,AI/AN,Asian,Black,Hispanic,Multi,NH/PI,Other,White,Unknown'
prepend_line('zipcode_vaccine.csv', column_labels)
