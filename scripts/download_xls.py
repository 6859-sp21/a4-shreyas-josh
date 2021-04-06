# Returns all Tuesdays and Thursdays of a given year
from datetime import datetime as dt
from datetime import date 
import calendar
import requests
from pandas import read_excel
from openpyxl import load_workbook

def get_thursdays():
    years = [2021]
    c = calendar.TextCalendar(calendar.SUNDAY)
    dates = []
    for year in years:
        for m in range(1,13):
            for i in c.itermonthdays(year,m):
                if i != 0:                                      #calendar constructs months with leading zeros (days belongng to the previous month)
                    day = date(year,m,i)
                    if day.weekday() == 3: 
                        dates.append("%s-%s-%s" % (calendar.month_name[m],i,year))
                    if date(year,m,i) == dt.today().date():
                        return dates

dates = get_thursdays()

prefix = "https://www.mass.gov/doc/weekly-covid-19-municipality-vaccination-report-"
postfix = "/download"
all_dfs = []

f = open("race_population_vaccine.csv", "w")

for d in dates:
    url = prefix + d.lower() + postfix
    resp = requests.get(url)
    content_type = resp.headers['Content-Type']
    if content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        fname = d + '.xlsx'
        with open(fname, 'wb') as r:
            r.write(resp.content)
        #my_sheet = 'Race and ethnicity - muni.' # change it to your sheet name, you can find your sheet name at the bottom left of your excel file
        wb = load_workbook(fname)
        ws = wb[wb.sheetnames[2]]
        for row in ws.iter_rows(min_row=2, max_col=6):
            row_str = d
            for cell in row:
                if str(cell.value) == 'County':
                    row_str = 'Date'
                row_str += ',' + str(cell.value) 
            row_str = row_str.strip(',')
            #print(row_str)    
            row_str += '\n'
            f.write(row_str)
        # for row in ws.values:
           # for value in row:
             # print(value)
    

        #df = read_excel(fname, sheet_name = my_sheet, engine = 'openpyxl')
        #dfs = read_excel(fname, sheet_name = None, engine = 'openpyxl')
f.close()
