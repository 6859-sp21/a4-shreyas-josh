# Returns all Tuesdays and Thursdays of a given year
from datetime import datetime as dt
from datetime import date 
import calendar
import requests
from pandas import read_excel
from openpyxl import load_workbook

f = open("../data/race_zipcode.csv", "w")

fname = '../data/zipcodes.xlsx'
#my_sheet = 'Race and ethnicity - muni.' # change it to your sheet name, you can find your sheet name at the bottom left of your excel file
wb = load_workbook(fname)
ws = wb[wb.sheetnames[1]]

row_str = ""
for row in ws.iter_rows(min_row=1, max_col=7):
    reset_row_str = True
    for cell in row:
        val = str(cell.value)
        if 'ZCTA5' in val and len(val) > 5:
            row_str = f'"{val.split(" ")[1]}"'
            reset_row_str = False
            break
        if val == "Estimate":
            continue
        if 'White' in val:
            val = 'White'
        if 'Black' in val:
            val = 'Black'
        if 'Indian' in val:
            val = 'AI/AN'
        if 'Asian' in val:
            val = 'Asian'
        if 'Pacific' in val:
            val = 'NH/PI'
        row_str += ',' + f'"{val}"'

    if reset_row_str:
        row_str = row_str.strip(',')
        row_str += '\n'
        f.write(row_str)
        row_str = ""
f.close()
